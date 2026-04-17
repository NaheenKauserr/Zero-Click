"""
report_generator.py — PDF + Excel report generation matching reference design.
"""
import io, os, datetime, warnings
import numpy as np, pandas as pd
warnings.filterwarnings("ignore")

try:
    from fpdf import FPDF
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False

try:
    import plotly.express as px
    import plotly.graph_objects as go
    HAS_PLOTLY = True
except ImportError:
    HAS_PLOTLY = False

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    HAS_OXL = True
except ImportError:
    HAS_OXL = False

NOW = datetime.datetime.now()

# ── helpers ──────────────────────────────────────────────────────────────
def _col_types(df):
    return {
        "numeric":     df.select_dtypes(include="number").columns.tolist(),
        "categorical": df.select_dtypes(include=["object","category"]).columns.tolist(),
        "datetime":    df.select_dtypes(include=["datetime64","datetimetz"]).columns.tolist(),
    }

def _s(df, ct):
    return next((c for c in ct["numeric"] if any(w in c.lower() for w in ["sales","revenue","amount"])),
                ct["numeric"][0] if ct["numeric"] else None)

def _p(df, ct):
    return next((c for c in ct["numeric"] if "profit" in c.lower()), None)

def _cat(df, ct):
    return next((c for c in ct["categorical"] if any(w in c.lower() for w in ["categ","type","segment"])), None)

def _reg(df, ct):
    return next((c for c in df.columns if "region" in c.lower()), None)

def _num(x):
    try: return float(x)
    except: return 0.0

def sales_by_region(df, ct):
    rc, sc = _reg(df, ct), _s(df, ct)
    if not rc or not sc:
        return pd.DataFrame({"Region":["N/A"],"Sales":[0]})
    out = df.groupby(rc)[sc].sum().reset_index().sort_values(sc, ascending=False)
    out.columns = ["Region","Sales"]
    return out.round(2)

def sales_by_category(df, ct):
    cc, sc = _cat(df, ct), _s(df, ct)
    if not cc or not sc:
        return pd.DataFrame({"Category":["N/A"],"Sales":[0],"Share %":[100]})
    out = df.groupby(cc)[sc].sum().reset_index().sort_values(sc, ascending=False)
    out.columns = ["Category","Sales"]
    tot = out["Sales"].sum()
    out["Share %"] = (out["Sales"]/tot*100).round(1) if tot else 0
    return out.round(2)

def top_products(df, ct, n=10):
    prod = next((c for c in df.columns if "product" in c.lower() and "name" in c.lower()), None)
    if not prod: prod = next((c for c in ct["categorical"] if "product" in c.lower()), None)
    sc, pc = _s(df, ct), _p(df, ct)
    if not prod or not sc:
        return pd.DataFrame({"Product":["N/A"],"Sales":[0]})
    gcols = [sc]+([pc] if pc else [])
    out = df.groupby(prod)[gcols].sum().reset_index().sort_values(sc, ascending=False).head(n)
    out.insert(0,"Rank",range(1,len(out)+1))
    cols = ["Rank","Product","Sales"]+( ["Profit"] if pc else [])
    out.columns = cols
    return out.round(2)

def monthly_trend(df, ct):
    dcs = ct.get("datetime",[])
    sc = _s(df, ct)
    if not dcs or not sc:
        return pd.DataFrame({"Date":[],"Sales":[],"MoM Growth":[]})
    dc = dcs[0]
    tmp = df.copy()
    tmp[dc] = pd.to_datetime(tmp[dc], errors="coerce")
    tmp["_p"] = tmp[dc].dt.to_period("M").dt.to_timestamp()
    m = tmp.groupby("_p")[sc].sum().reset_index()
    m.columns = ["Date","Sales"]
    m["MoM Growth"] = m["Sales"].pct_change().fillna(0).round(4)
    return m.round(2)

# ════════════════════════════════════════════════════════════════════════
# PDF
# ════════════════════════════════════════════════════════════════════════
if HAS_FPDF:
    class GenesisReport(FPDF):
        def header(self):
            self.set_fill_color(124,58,237); self.rect(0,0,210,18,"F")
            self.set_font("helvetica","B",12); self.set_text_color(255,255,255)
            self.set_y(3); self.cell(0,12,"Zero Click AI  —  Comprehensive Analysis Report",align="C")
            self.set_y(18); self.set_fill_color(167,139,250); self.rect(0,18,210,6,"F")
            self.set_font("helvetica","",7); self.set_y(19)
            self.cell(100,4,f"  Report Date: {NOW.strftime('%Y-%m-%d')}  |  Genesis Training",align="L")
            self.cell(0,4,"Zero Click AI  |  AI-Powered Analytics  ",align="R"); self.ln(10)

        def footer(self):
            self.set_y(-13); self.set_font("helvetica","I",7); self.set_text_color(150,150,150)
            self.cell(0,8,f"Page {self.page_no()}  |  Generated {NOW.strftime('%Y-%m-%d %H:%M')}  |  Genesis Training",align="C")

        def section_title(self,text):
            self.ln(5); self.set_font("helvetica","B",12); self.set_text_color(124,58,237)
            self.cell(0,8,text,ln=True); self.set_draw_color(167,139,250); self.set_line_width(0.4)
            self.line(10,self.get_y(),200,self.get_y()); self.ln(3)
            self.set_text_color(30,30,30); self.set_font("helvetica","",9)

        def kpi_box(self,label,value,delta,positive=True):
            x,y,w=self.get_x(),self.get_y(),43
            self.set_fill_color(245,243,255); self.rect(x,y,w,18,"F")
            self.set_draw_color(167,139,250); self.rect(x,y,w,18)
            self.set_font("helvetica","B",13)
            col=(16,185,129) if positive else (239,68,68); self.set_text_color(*col)
            self.set_xy(x,y+1); self.cell(w,7,value,align="C")
            sym="▲" if positive else "▼"; self.set_font("helvetica","",8)
            self.set_xy(x,y+8); self.cell(w,5,f"{sym} {delta}",align="C")
            self.set_font("helvetica","",7); self.set_text_color(100,100,100)
            self.set_xy(x,y+13); self.cell(w,4,label,align="C")
            self.set_xy(x+w+2,y)

        def mini_table(self,headers,rows,col_widths,rh=6):
            self.set_fill_color(124,58,237); self.set_font("helvetica","B",8); self.set_text_color(255,255,255)
            for h,w in zip(headers,col_widths): self.cell(w,7,str(h)[:20],border=1,fill=True,align="C")
            self.ln()
            self.set_font("helvetica","",8)
            for ri,row in enumerate(rows):
                self.set_fill_color(250,248,255) if ri%2==0 else self.set_fill_color(255,255,255)
                self.set_text_color(40,40,40)
                for val,w in zip(row,col_widths): self.cell(w,rh,str(val)[:22],border=1,fill=True,align="C")
                self.ln()

        def write_ai(self,text):
            self.set_font("helvetica","",9); self.set_text_color(40,40,40)
            for line in text.split("\n"):
                line=line.strip()
                if not line: self.ln(2); continue
                if line.startswith("## "):
                    self.ln(3); self.set_font("helvetica","B",10); self.set_text_color(124,58,237)
                    self.multi_cell(0,6,line[3:]); self.set_font("helvetica","",9); self.set_text_color(40,40,40)
                elif line.startswith(("- ","• ","* ")):
                    self.multi_cell(0,5,f"  • {line[2:]}")
                else:
                    self.multi_cell(0,5,line)


def generate_report_pdf(df, kpis, ml_results, forecast_data, insights, charts) -> bytes:
    if not HAS_FPDF:
        return b""
    ct = _col_types(df)
    pdf = GenesisReport()
    pdf.set_auto_page_break(auto=True, margin=20)
    pdf.set_margins(10, 28, 10)

    # ── Page 1: Performance Dashboard ─────────────────────────────────
    pdf.add_page()
    pdf.set_fill_color(240,237,255); pdf.rect(10,28,190,12,"F")
    pdf.set_font("helvetica","B",13); pdf.set_text_color(124,58,237)
    pdf.set_y(30); pdf.cell(0,8,"PERFORMANCE DASHBOARD",align="C",ln=True)

    kpi_list = list(kpis or [])
    while len(kpi_list) < 6: kpi_list.append(("N/A","—","—"))
    pdf.set_y(44); pdf.set_x(11)
    for lbl,val,dlt in kpi_list[:3]: pdf.kpi_box(str(lbl)[:18],str(val)[:12],str(dlt)[:10] if len(kpi_list[0])>2 else "—",True)
    pdf.ln(20); pdf.set_x(11)
    for lbl,val,dlt in kpi_list[3:6]: pdf.kpi_box(str(lbl)[:18],str(val)[:12],str(dlt)[:10] if len(kpi_list[0])>2 else "—",True)
    pdf.ln(22)

    ai_y=pdf.get_y(); pdf.set_fill_color(248,248,255); pdf.set_draw_color(167,139,250)
    pdf.rect(10,ai_y,190,5,"F"); pdf.set_font("helvetica","B",9); pdf.set_text_color(124,58,237)
    pdf.cell(0,5,"  AI NARRATED SUMMARY",ln=True)
    pdf.set_font("helvetica","",9); pdf.set_text_color(40,40,40)
    ai_short=""
    if isinstance(insights,str):
        lines=[l.strip() for l in insights.split("\n") if l.strip() and not l.startswith("#")]
        ai_short=" ".join(lines[:3])[:380]
    elif isinstance(insights,list):
        ai_short=" ".join(str(i) for i in insights[:3])[:380]
    pdf.multi_cell(190,5,ai_short or "AI analysis complete.")
    pdf.set_font("helvetica","I",7); pdf.set_text_color(120,120,120)
    pdf.cell(0,4,"*Profit margin = Profit/Sales",ln=True); pdf.ln(3)

    # ── Build chart images ─────────────────────────────────────────────
    tmp = "/tmp/genesis_pdf_imgs"; os.makedirs(tmp,exist_ok=True)
    chart_paths=[]

    if HAS_PLOTLY:
        extra=[]
        try:
            rbr=sales_by_region(df,ct)
            fig=px.bar(rbr,x="Sales",y="Region",orientation="h",
                       color="Sales",color_continuous_scale="Purples",title="Sales by Region")
            fig.update_layout(paper_bgcolor="white",plot_bgcolor="white",font_color="#222",
                              margin=dict(l=10,r=10,t=40,b=10),height=280,coloraxis_showscale=False)
            extra.append(("Sales by Region",fig))
        except: pass

        try:
            sbc=sales_by_category(df,ct)
            fig=px.pie(sbc,names="Category",values="Sales",hole=0.45,
                       title="Sales by Category",color_discrete_sequence=px.colors.sequential.Purples_r)
            fig.update_traces(textinfo="percent+label",textfont_size=9)
            fig.update_layout(paper_bgcolor="white",font_color="#222",
                              margin=dict(l=10,r=10,t=40,b=10),height=280)
            extra.append(("Sales by Category",fig))
        except: pass

        try:
            mt=monthly_trend(df,ct)
            if not mt.empty:
                fig=px.line(mt,x="Date",y="Sales",title="Monthly Sales Trend",color_discrete_sequence=["#7c3aed"])
                fig.update_traces(fill="tozeroy",fillcolor="rgba(124,58,237,0.08)")
                fig.update_layout(paper_bgcolor="white",plot_bgcolor="white",font_color="#222",
                                  margin=dict(l=10,r=10,t=40,b=10),height=280)
                extra.append(("Monthly Sales Trend",fig))
        except: pass

        try:
            sc,pc=_s(df,ct),_p(df,ct)
            if sc and pc:
                scat=df[[sc,pc]].dropna().sample(min(300,len(df)),random_state=42)
                fig=px.scatter(scat,x=sc,y=pc,title="Sales vs Profit",color_discrete_sequence=["#7c3aed"])
                fig.update_layout(paper_bgcolor="white",plot_bgcolor="white",font_color="#222",
                                  margin=dict(l=10,r=10,t=40,b=10),height=280)
                extra.append(("Sales vs Profit",fig))
        except: pass

        all_c = extra + list(charts or [])
        for i,(title,fig) in enumerate(all_c[:8]):
            try:
                p=os.path.join(tmp,f"img_{i}.png")
                fig.write_image(p,engine="kaleido",width=600,height=300,scale=2)
                chart_paths.append((title,p))
            except Exception as e: print(f"chart img {i}: {e}")

    # 2 per page, side by side
    for i in range(0,len(chart_paths),2):
        pdf.add_page()
        base_y = pdf.get_y()
        for j in range(2):
            if i+j >= len(chart_paths): break
            title,path=chart_paths[i+j]
            x_pos=10 if j==0 else 108
            try:
                pdf.set_font("helvetica","B",9); pdf.set_text_color(80,80,80)
                pdf.set_xy(x_pos, base_y)
                pdf.cell(90,6,title,align="C")
                pdf.image(path,x=x_pos,y=base_y+7,w=90)
            except: pass
        pdf.ln(60)

    # ── Top Products ──────────────────────────────────────────────────
    pdf.add_page(); pdf.section_title("Top Products by Sales & Profit")
    tp=top_products(df,ct,n=16)
    if not tp.empty:
        hdrs=list(tp.columns); cw=[10]+[int(160/(len(hdrs)-1))]*(len(hdrs)-1)
        rows=[list(tp.iloc[i]) for i in range(len(tp))]
        pdf.mini_table(hdrs,rows,cw)

    # ── ML Insights ───────────────────────────────────────────────────
    pdf.add_page(); pdf.section_title("Machine Learning Insights")
    if ml_results:
        for res in ml_results:
            pdf.set_font("helvetica","B",10); pdf.set_text_color(124,58,237)
            pdf.cell(0,7,f"  {res.get('title','Model')}",ln=True)
            pdf.set_font("helvetica","",9); pdf.set_text_color(0,130,70)
            pdf.cell(0,5,f"    {res.get('metric_name')}: {res.get('metric_value')}",ln=True)
            pdf.set_text_color(60,60,60)
            pdf.multi_cell(0,5,f"    Type: {res.get('type')} | {res.get('extra','')}"); pdf.ln(3)
    else:
        pdf.set_text_color(100,100,100)
        pdf.cell(0,6,"  ML models not trained — need numeric columns.",ln=True)

    # ── Forecast ──────────────────────────────────────────────────────
    if forecast_data:
        pdf.add_page(); pdf.section_title("Time-Series Forecast")
        pdf.set_font("helvetica","",9); pdf.set_text_color(60,60,60)
        pdf.cell(0,5,f"  Model: {forecast_data.get('model','N/A')}  |  6-month horizon",ln=True); pdf.ln(3)
        fdf=forecast_data.get("forecast_df")
        if fdf is not None and not fdf.empty:
            hdrs=["Period","Predicted","Lower 95%","Upper 95%"]; cw=[50,45,45,45]
            rows=[]
            for _,row in fdf.head(12).iterrows():
                dt=str(row.get("ds","")).split(" ")[0]
                rows.append([dt,f"${_num(row.get('yhat',0)):,.0f}",
                              f"${_num(row.get('yhat_lower',0)):,.0f}",
                              f"${_num(row.get('yhat_upper',0)):,.0f}"])
            pdf.mini_table(hdrs,rows,cw)

    # ── Full AI Insights ──────────────────────────────────────────────
    pdf.add_page(); pdf.section_title("Full AI Insights")
    if insights: pdf.write_ai(insights if isinstance(insights,str) else "\n".join(f"- {i}" for i in insights))

    # ── Data Sample ───────────────────────────────────────────────────
    pdf.add_page(); pdf.section_title("Dataset Sample — First 20 Rows")
    sample=df.head(20).astype(str); cols=list(sample.columns[:7])
    cw=[int(185/len(cols))]*len(cols)
    rows=[[str(sample.iloc[i][c])[:18] for c in cols] for i in range(len(sample))]
    pdf.mini_table([c[:16] for c in cols],rows,cw,rh=5)

    try:
        for f in os.listdir(tmp): os.remove(os.path.join(tmp,f))
        os.rmdir(tmp)
    except: pass

    return bytes(pdf.output())


# ════════════════════════════════════════════════════════════════════════
# EXCEL
# ════════════════════════════════════════════════════════════════════════

def _hdr(ws, row, ncols, bg="7C3AED", fg="FFFFFF"):
    fill=PatternFill("solid",fgColor=bg); fnt=Font(bold=True,color=fg,size=10)
    bdr=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for c in range(1,ncols+1):
        cell=ws.cell(row=row,column=c)
        cell.fill=fill; cell.font=fnt; cell.border=bdr
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

def _row(ws, ri, vals, alt=False):
    bg="F5F3FF" if alt else "FFFFFF"
    fill=PatternFill("solid",fgColor=bg)
    bdr=Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))
    for ci,val in enumerate(vals,1):
        cell=ws.cell(row=ri,column=ci,value=val)
        cell.fill=fill; cell.border=bdr; cell.font=Font(size=9)
        cell.alignment=Alignment(horizontal="center",vertical="center")

def _write_df(ws, df_in, start=2):
    hdrs=list(df_in.columns); _hdr(ws,start,len(hdrs))
    for ci,h in enumerate(hdrs,1):
        ws.cell(row=start,column=ci,value=h)
        ws.column_dimensions[get_column_letter(ci)].width=max(14,len(str(h))+4)
    for ri,(_, row) in enumerate(df_in.iterrows(),1):
        vals=list(row); _row(ws,start+ri,vals,alt=ri%2==0)
        for ci,val in enumerate(vals,1):
            cell=ws.cell(row=start+ri,column=ci)
            cn=hdrs[ci-1].lower()
            if "mom" in cn or "growth" in cn:
                try:
                    v=float(val)
                    cell.font=Font(size=9,bold=True,color="27500A" if v>=0 else "A32D2D")
                    cell.value=f"+{v*100:.2f}%" if v>=0 else f"{v*100:.2f}%"
                except: pass
            elif any(w in cn for w in ["sales","revenue","profit","amount"]):
                try: cell.number_format="#,##0.00"; cell.value=float(val)
                except: pass

def _title(ws,text,row=1,end_col=8):
    ws.merge_cells(start_row=row,start_column=1,end_row=row,end_column=end_col)
    c=ws.cell(row=row,column=1,value=text)
    c.fill=PatternFill("solid",fgColor="7C3AED"); c.font=Font(bold=True,color="FFFFFF",size=12)
    c.alignment=Alignment(horizontal="center",vertical="center"); ws.row_dimensions[row].height=22

def generate_report_excel(df, kpis, ml_results, forecast_data, insights) -> bytes:
    if not HAS_OXL:
        return b""
    ct=_col_types(df)
    wb=openpyxl.Workbook()

    # README
    ws=wb.active; ws.title="README"
    _title(ws,"Zero Click AI — Analysis Report",row=1,end_col=6)
    info=[("Report Date",NOW.strftime("%Y-%m-%d %H:%M")),("Platform","Zero Click AI · Genesis Training"),
          ("Dataset Rows",f"{len(df):,}"),("Dataset Cols",f"{len(df.columns)}"),
          ("AI Insights",(insights[:200] if isinstance(insights,str) else str(insights)[:200]))]
    for ri,(k,v) in enumerate(info,3):
        ws.cell(row=ri,column=1,value=k).font=Font(bold=True,color="7C3AED",size=10)
        ws.cell(row=ri,column=2,value=v).font=Font(size=10)
        ws.row_dimensions[ri].height=16
    ws.column_dimensions["A"].width=20; ws.column_dimensions["B"].width=60

    # KPIs
    ws2=wb.create_sheet("KPIs"); _title(ws2,"Key Performance Indicators",end_col=4)
    kpi_rows=[[str(k[0]),str(k[1]),str(k[2]) if len(k)>2 else ""] for k in (kpis or [])]
    _hdr(ws2,2,3); [ws2.cell(row=2,column=c+1,value=h) for c,h in enumerate(["Metric","Value","Description"])]
    [_row(ws2,i+3,r,alt=i%2==0) for i,r in enumerate(kpi_rows)]
    ws2.column_dimensions["A"].width=30; ws2.column_dimensions["B"].width=20; ws2.column_dimensions["C"].width=40
    if len(kpi_rows)>=2:
        ch=BarChart(); ch.title="KPI Values"; ch.style=10; ch.width=16; ch.height=9
        data=Reference(ws2,min_col=2,min_row=2,max_row=2+len(kpi_rows))
        cats=Reference(ws2,min_col=1,min_row=3,max_row=2+len(kpi_rows))
        ch.add_data(data,titles_from_data=True); ch.set_categories(cats)
        ws2.add_chart(ch,"E2")

    # Sales by Region
    ws3=wb.create_sheet("Sales_by_Region"); _title(ws3,"Sales by Region",end_col=4)
    rbr=sales_by_region(df,ct); _write_df(ws3,rbr,start=2)
    if len(rbr)>=2:
        ch=BarChart(); ch.title="Sales by Region"; ch.style=10; ch.width=16; ch.height=9
        data=Reference(ws3,min_col=2,min_row=2,max_row=2+len(rbr))
        cats=Reference(ws3,min_col=1,min_row=3,max_row=2+len(rbr))
        ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ws3.add_chart(ch,"D2")

    # Sales by Category
    ws4=wb.create_sheet("Sales_by_Category"); _title(ws4,"Sales by Category",end_col=5)
    sbc=sales_by_category(df,ct); _write_df(ws4,sbc,start=2)
    if len(sbc)>=2:
        ch=PieChart(); ch.title="Category Distribution"; ch.style=10; ch.width=14; ch.height=9
        data=Reference(ws4,min_col=2,min_row=2,max_row=2+len(sbc))
        cats=Reference(ws4,min_col=1,min_row=3,max_row=2+len(sbc))
        ch.add_data(data,titles_from_data=True); ch.set_categories(cats); ws4.add_chart(ch,"E2")

    # Top 10 Products
    ws5=wb.create_sheet("Top10_Products_Profit"); _title(ws5,"Top 10 Products by Sales & Profit",end_col=6)
    tp=top_products(df,ct,n=10); _write_df(ws5,tp,start=2)
    if "Profit" in tp.columns:
        pci=list(tp.columns).index("Profit")+1
        for ri in range(3,3+len(tp)):
            cell=ws5.cell(row=ri,column=pci)
            try:
                v=float(cell.value)
                cell.fill=PatternFill("solid",fgColor="EAFAF1" if v>0 else "FDEDEC")
                cell.font=Font(color="27500A" if v>0 else "A32D2D",size=9,bold=True)
            except: pass

    # Monthly Trend
    ws6=wb.create_sheet("Monthly_Trend"); _title(ws6,"Monthly Sales Trend",end_col=5)
    mt=monthly_trend(df,ct)
    if mt.empty: ws6.cell(row=3,column=1,value="No datetime column found.")
    else:
        _write_df(ws6,mt,start=2)
        ch=LineChart(); ch.title="Monthly Sales Trend"; ch.style=10; ch.width=16; ch.height=9
        data=Reference(ws6,min_col=2,min_row=2,max_row=2+len(mt))
        ch.add_data(data,titles_from_data=True); ws6.add_chart(ch,"E2")

    # Raw Data
    ws7=wb.create_sheet("Raw_Data"); end=min(len(df.columns),10)
    _title(ws7,"Raw Cleaned Dataset",end_col=end); _write_df(ws7,df.head(500),start=2)

    for ws in [ws2,ws3,ws4,ws5,ws6,ws7]: ws.freeze_panes="A3"

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf.read()
